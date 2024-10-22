import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def log_data(directory, row_data):
    """Create a new Excel file or append to an existing one based on the date."""
    
    # Generate a filename based on the current month and year
    now = datetime.now()
    filename = f"Tools In-Out {now.strftime('%B %Y')}.xlsx"
    full_path = os.path.join(directory, filename)
    
    headers = ['Date', 'Time', 'Operator Name', 'Machine Code', 'Tool Type', 'Tool Size', 'Quantity Issued', 'Quantity Received']

    # Check if the file exists
    if not os.path.exists(full_path):
        # Create a new workbook, add headers, and set column width
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for i, header in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(i)].width = 20
    else:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(full_path)
        sheet = workbook.active

    # Get current date and time
    current_date = now.strftime('%Y-%m-%d')
    current_time = now.strftime('%I:%M:%S %p')

    # Prepend date and time to the row data
    full_row_data = [current_date, current_time] + row_data

    # Append the data
    sheet.append(full_row_data)
    workbook.save(full_path)
    print(f"Data saved to {full_path}")